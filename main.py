'''
File: main.py
Author: Alvin La
Description: This file obtains data from the 42i-TL and 49C Thermo Scientific Instrument. Spectifiically
NoX No2 No from 42i-TL and O3 from 49C. Datalogs every one second.
'''
import serial_func as sf
from datetime import datetime as dt, timedelta as td
from openpyxl import Workbook
instrument_serial_info = {"42i-TL" : {"com_port": "",
                                      "baudrate" : 0,
                                      "id" : chr(128 + 42),
                                      "avaliable_baudrate": [1200, 2400 , 4800, 9600 , 19200 , 38400 , 57600 , 115200],
                                      "measurements" : ["no", "no2", "nox"],
                                      "serial": None},
                          "49C":  {"com_port": "",
                                   "baudrate" : 0,
                                   "id" : chr(128 + 49),
                                   "avaliable_baudrate" : [1200, 2400 , 4800, 9600],
                                   "measurements" : ["o3"],
                                   "serial": None}}
def main():
    #gets computer's avaliable com ports
    com_list = sf.get_com_ports()
    for instrument in instrument_serial_info.keys():
        valid_com_port = False
        valid_baudrate = False
        while not valid_com_port:
            print("\nList of available COM ports:")
            for port in com_list:
                print(port)
            com_input = input(f"Select COM port for {instrument}: ").upper()
            if com_input not in com_list:
                print("unknown COM port\n")
            else:
                valid_com_port = True
        while not valid_baudrate:
            print("\nList of available Baudrate:")
            for baudrate in instrument_serial_info[instrument]["avaliable_baudrate"]:
                print(baudrate)
            baurdrate_input = input(f"Select Baudrate for {instrument}: ")
            if baurdrate_input.isdigit():
                baurdrate_input = int(baurdrate_input)
                if baurdrate_input not in instrument_serial_info[instrument]["avaliable_baudrate"]:
                    print("Unknown baudrate")
                else:
                    valid_baudrate = True
            else:
                print("Invalid Input")
        instrument_serial_info[instrument]["com_port"] = com_input
        instrument_serial_info[instrument]["baudrate"] = baurdrate_input
        com_list.remove(com_input)
    print()
    for instrument in instrument_serial_info:
        print(instrument + ":", instrument_serial_info[instrument]["com_port"], instrument_serial_info[instrument]['baudrate'])
    for instrument in instrument_serial_info:
        ser_temp = sf.open_serial_port(instrument_serial_info[instrument]['com_port'], instrument_serial_info[instrument]["baudrate"])
        if not ser_temp:
            print("Unable to open Serial port for", instrument)
            for instrument in instrument_serial_info:
                if instrument_serial_info[instrument]["serial"]:
                    instrument_serial_info[instrument]["serial"].close()
            return
        else:
            instrument_serial_info[instrument]["serial"] = ser_temp
    start_time = dt.now()
    second_timer = dt.now()
    file_name = f'{start_time.date()}_{start_time.hour}-{start_time.minute}-{start_time.second}.xlsx'
    print(file_name)
    instrument_data = {
        "42i-TL" : [],
        "49C" : []
    }
    
    #checks if 3 hours has passed since sampling duration is 3 hours
    try:
        while dt.now() < start_time + td(seconds=15):
            #checks if a second has passed
            if dt.now() - second_timer >= td(seconds=1):
                data_collection_time = dt.now()
                for instrument in instrument_serial_info:
                    #writes to the instrument requesting for the data
                    data = sf.get_data(instrument_serial_info[instrument]["serial"], instrument_serial_info[instrument]["measurements"], instrument_serial_info[instrument]["id"])
                    list_data = [f'{data_collection_time}']
                    for measurements in data:
                        list_data.append(data[measurements]['value'])
                    #creates first row of headers
                    if len(instrument_data[instrument]) == 0:
                        headers = ["Time"]
                        for measurements in data:
                            headers.append(f'{measurements} ({data[measurements]["unit"]})')
                        instrument_data[instrument].append(headers)    
                    instrument_data[instrument].append(list_data)
                    print(list_data)
                second_timer = dt.now()
                print()
    except KeyboardInterrupt:
        print("Saving Data and Exiting Program")
        print(f'Duration: {dt.now() - start_time}')
    wb = Workbook()
    #creates two sheets
    data_ws = {
        "42i-TL" : wb.active,
        "49C" : wb.create_sheet()
    }
    for instrument in instrument_data:
        #names the sheet after the instrument
        data_ws[instrument].title = instrument
        #loops through each data point and prints it to the excel cell
        for row_idx, row_data in enumerate(instrument_data[instrument], start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                data_ws[instrument].cell(row=row_idx, column=col_idx, value=cell_value)
    wb.save(file_name)
    for instrument in instrument_serial_info:
        instrument_serial_info[instrument]["serial"].close()
        instrument_serial_info[instrument]['serial'] = None
def excel_test():
    wb = Workbook()
    ws = {
        "ws1" : wb.active,
        "ws2" : wb.create_sheet()
    }
    for worksheet in ws:
        ws[worksheet].title = worksheet
    data1 = [["Column1", "Column2", "Column3"],
             [0, 1, 2],
             [3, 4, 5],
             [6, 7, 8],
             [9, 10, 11],
             [12, 13, 14],
             [15, 16, 17]]
    data2 = [["Column1", "Column2"],
             [0, 1],
             [2, 3],
             [4, 5],
             [6, 7],
             [8, 9],
             [10, 11]]
    for row_idx, row_data in enumerate(data1, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            ws['ws1'].cell(row=row_idx, column=col_idx, value=cell_value)
    for row_idx, row_data in enumerate(data2, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            ws['ws2'].cell(row=row_idx, column=col_idx, value=cell_value)
    wb.save("testfilewithdata.xlsx")

def serial_test():
    ser_42i = sf.open_serial_port("COM8", 9600)
    if not ser_42i:
        print("Unable to open 42iport")
        return
    print(sf.get_data(ser_42i, instrument_serial_info["42i-TL"]['measurements'], chr(170)))    
    ser_49C = sf.open_serial_port("COM9", 9600)
    if not ser_49C:
        ser_49C.close()
        print("Unable to open 49C port")
        return
    print(sf.get_data(ser_49C, instrument_serial_info["49C"]["measurements"],  chr(128 + 49)))
    ser_49C.write(f'{chr(128 + 49)}o3\r'.encode())
    data = ser_49C.read_until(b'\r')
    print(data)

if __name__ == "__main__":
    #excel_test()
    main()
    #serial_test()

